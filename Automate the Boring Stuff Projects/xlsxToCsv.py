#! /usr/bin/env python3
# xlsxToCsv.py - Converts an Excel file to a CSV file

import openpyxl, csv, os

os.chdir('./materials/excelSpreadsheets')

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object
    if not excelFile.endswith('.xlsx'):
        continue # jumps back to loop if the filename does not end with xlsx
    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.sheetnames: # deprecated from get_sheet_names()
        # Loop through every sheet in the workbook.
        for sheet in sheetName:
            sheet = wb[sheetName]
            # Create the CSV filename from the Excel filename and sheet title.
            outputCsv = open(excelFile[:-5] + '.csv', 'w', newline='')
            
            # Create the csv.writer object for this CSV file.
            csvWriter = csv.writer(outputCsv)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = [] # append each cell to this initialized list
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData
                    cell = sheet.cell(row=rowNum, column=colNum)
                    rowData.append(cell.value)

                # Write the rowData list to the CSV file.
                csvWriter.writerow(rowData)

            outputCsv.close()
